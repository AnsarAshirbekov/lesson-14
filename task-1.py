# а) Прочитайте из трёх excel файлов (заранее создайте их, внутри 1111, 2222, 3333).

# В условии задачи сказано прочитать 3 РАЗНЫХ excel-файла, внутри которых написаны значения
# 1111, 2222, 3333 соответственно. Но тут не уточнили, что эти значения должны быть в одной 
# ячейке excel-файла или в разных, в строку или в столбец.
# Будем считать что в разных ячейках в строку. 

# Импортируем из модуля openpyxl функции для создания и прочтения excel-файлов
from openpyxl import Workbook, load_workbook

# Для начала создадим эти 3 файла с помощью функции Workbook(), которая просто создает новую пустую 
# рабочую книгу в Excel
workbook1 = Workbook()

# Далее создаем активный лист в этой рабочей книге, с которым мы будем взаимодействовать. Как мы 
# помним каждый excel-файл (рабочая книга) содержит в себе листы (Sheet).
work_sheet1 = workbook1.active

# Теперь с помощью цикла записываем в этот лист значения 1 в каждую ячейку первой строки
for index in range(1, 5): # так как нам сказано записать 4 значения (1111) значит создаем точный диапазон
     cell = f'{chr(64+index)}1' # Здесь нам надо указать номер ячейки первой строки, то есть А1, В1, С1, D1.
                                # Как мы видим тут меняются только буквы, а цикл будет перебирать только числа.
                                # Значит конвертируем числа в буквы с помощью таблицы ASCII. 
                                # A, B, C, D в таблице ASCII соответствуют числам 65, 66, 67, 68.
                                # Цикл будет перебирать числа 1, 2, 3, 4 из диапазона range.
                                # Значит чтобы получить числа для букв мы в каждой итерации прибавляем 64 к каждому
                                # итерируемому элементу и с помощью функции chr() конвертируем число в букву
     work_sheet1[cell] = '1' # Обращаемся к соответствующей ячейке листа и записываем туда значения 1

# Сохраняем нашу рабочую книгу, указывая путь
workbook1.save('Completed Homeworks Python\\lesson 14\\test1.xlsx')

# Проделываем все тоже самое для 2-ой и 3-й рабочих книг (excel-файлов)
workbook2 = Workbook()

work_sheet2 = workbook2.active

for index in range(1, 5):
    cell = f'{chr(64+index)}1'
    work_sheet2[cell] = '2'

workbook2.save('Completed Homeworks Python\\lesson 14\\test2.xlsx')

workbook3 = Workbook()

work_sheet3 = workbook3.active

for index in range(1, 5):
    cell = f'{chr(64+index)}1'
    work_sheet3[cell] = '3'

workbook3.save('Completed Homeworks Python\\lesson 14\\test3.xlsx')

# Теперь нам нужно прочитать то, что мы записали в эти файлы и вывести в консоль
# Для этого мы используем функцию .load_workbook модуля openpyxl, который принимает excel-файл, находящийся в 
# в определенном пути, и создает объект рабочей книги
workbook1_read = load_workbook('Completed Homeworks Python\\lesson 14\\test1.xlsx')

# Активируем рабочий лист этой книги
work_sheet1_read = workbook1_read.active

# Создаем пустую строку для сохранения в нее значений из рабочего листа excel-файла
text1 = ''

# Так как мы точно знаем, что в этом листе значения находятся в первых четырех ячейках первой строки, то создаем
# похожий цикл для прочтения на тот, который у нас был при записи
for index in range(1, 5):
    # Здесь мы обращаемся к ячейке рабочего листа, вытаскиваем из него значение с помощью метода .value, конвертируем 
    # все это в строку и производим операцию конкатенации для всех полученных значений к заранее созданной пустой строке
    text1 += str(work_sheet1_read[f'{chr(64+index)}1'].value)

# Выводим полученную строку в консоль
print(text1)

# Тоже самое проделываем для остальных файлов 
workbook2_read = load_workbook('Completed Homeworks Python\\lesson 14\\test2.xlsx')

work_sheet2_read = workbook2_read.active

text2 = ''

for index in range(1, 5):
    text2 += str(work_sheet2_read[f'{chr(64+index)}1'].value)

print(text2)

workbook3_read = load_workbook('Completed Homeworks Python\\lesson 14\\test3.xlsx')

work_sheet3_read = workbook3_read.active

text3 = ''

for index in range(1, 5):
    text3 += str(work_sheet3_read[f'{chr(64+index)}1'].value)

print(text3)


# б) Отсортируйте полученную матрицу в порядке убывания.

# Не знаю каким образом у меня должна была получить матрица, когда там было четко написано про 3 РАЗНЫХ excel-файла.
# Но да ладно, у меня есть 3 строки text1, text2, text3 со значениями '1111', '2222', '3333' соответственно.
# Мы можем легко сделать из них матрицу.

# Для этого сначала преобразуем строки в числа, создав списки с числами
text1 = [int(i) for i in text1]

text2 = [int(i) for i in text2]

text3 = [int(i) for i in text3]

# Затем эти списки закинем в другой список, который будет матрицей (списком списков)
matrix = [text1, text2, text3]

# Отсортируем эту матрицу в порядке убывания
sorted_matrix = sorted(matrix, reverse=True)

print(sorted_matrix)


# в) Запишите это в один файл, изменив шрифт и обернув в границы.

workbook = Workbook()

work_sheet = workbook.active

# Здесь уже записывать значения в ячейки как в прошлый раз не выйдет (по крайней мере у меня)
# Нужно воспользоваться другими функциями, такой как например enumerate()
# Это функция используется для перебора элементов последовательностей (списки, строки, кортежи)
# Она принимает какой-то итерируемый объект и start, который указывает на начало отчета индексов элементов
# этого объекта. То есть start это начальное значение индекса. Он не обязателен, но по умолчанию равен 0.
# Функция возвращает индексы элементов и их значения. Основное применение этой функции - циклы for.
# Мы здесь указываем после for две итерируемые переменные: первая та, которая будет указывать на индекс элемента,
# вторая - на значение этого элемента. 
# Мы помним что в нашей матрице внутренние списки содержат числа, которые должны быть записаны в строку
# Значит сначала проходимся по строкам 
for row_index, row in enumerate(sorted_matrix, start=1): # здесь если бы хотели вывести все в консоль print(row_index, row)
                                                         # у нас бы вышло в каждой итерации:
                                                         # 1) row_index = 1; row = [3, 3, 3, 3]  
                                                         # 2) row_index = 2; row = [2, 2, 2, 2]  
                                                         # 3) row_index = 3; row = [1, 1, 1, 1]  
    # внутри проходимся по столбцам
    for column_index, value in enumerate(row, start=1): # print(column_index, value):
                                                          # 1)
                                                              # 1.1) column_index = 1; value = 3
                                                              # 1.2) column_index = 2; value = 3
                                                              # 1.3) column_index = 3; value = 3
                                                              # 1.4) column_index = 4; value = 3
                                                          # 2)
                                                              # 2.1) column_index = 1; value = 2
                                                              # 2.2) column_index = 2; value = 2
                                                              # 2.3) column_index = 3; value = 2
                                                              # 2.4) column_index = 4; value = 2
                                                          # 3)
                                                              # 3.1) column_index = 1; value = 1
                                                              # 3.2) column_index = 2; value = 1
                                                              # 3.3) column_index = 3; value = 1
                                                              # 3.4) column_index = 4; value = 1
        # далее мы используем метод .cell() который позволяет получить доступ к любой ячейке рабочего листа excel-файла
        # по заданным координатам row и column, которые являются обязательными для работы этого метода.
        # Чтобы ввести какое-нибудь значение в ячейку, нужно указать помимо координат и само значение
        work_sheet.cell(row=row_index, column=column_index, value=value)
        # Но если мы допустим хотим узнать значение какой-либо ячейки, то делаем следующим образом:
        # cell_A1_value = work_sheet.cell(row=1, column=1).value
        # здесь мы узнали значение ячейки А1, координаты которой строка = 1, столбец = 1 и записали в переменную

# Чтобы изменить шрифт и обернуть границы нужно импортировать из подмодуля openpyxl.styles следующие функции 
from openpyxl.styles import Font, Border, Side

# Создаем переменную для сохранения в нее информации о шрифте
font = Font(size=14) # Данная функция принимает не только размер шрифта, но также стили, жирность, цвет, курсив и т.д.

# Чтобы сохранить информацию о границах используем функцию Border(), которая принимает какую-либо сторону ячейки, будь то
# справа, слева, диагональ и т.д. Но Border() не может работать без Side(), то есть другой, функции которая принимает
# цвет, стиль и толщину для линии, которые мы хотим отредактировать
border = Border(left=Side(border_style='thin', color='FF0000'),
                right=Side(border_style='thin', color='FF0000'),
                top=Side(border_style='thin', color='FF0000'),
                bottom=Side(border_style='thin', color='FF0000')) # Здесь мы указали цвет (красный), и стиль для линий 
                                                                  # (иначе ничего не будет видно) которые будут 
                                                                  # являться границами ячейки

# Теперь нам нужно задать все вышеуказанные свойства для ячеек рабочего листа excel-файла
# Для этого используем метод .iter_rows(), который проходится по определенному диапазону ячеек рабочего листа и что-нибудь
# с ними делает. Извлекает значения, записывает значения и т.д. Он принимает координаты диапазона ячеек, по которым ему
# нужно пройтись, но если ему ничего не указывать, то есть (), то он пройдется по всем ячейкам рабочего листа
for row in work_sheet.iter_rows(): # Сначала он начнет проходится по строкам
    for cell in row: # а затем по ячейкам в этих строках
        cell.font = font # и запишет в них информацию о шрифте
        cell.border = border # и о границах

workbook.save('Completed Homeworks Python\\lesson 14\\sorted matrix.xlsx')